#!/usr/bin/env python3
"""
update_carousel_prompts.py

Updates the "Kie.ai Image Description" column in posts_schedule.xlsx for all
16 carousel posts with explicit, art-directed slide prompts.

Each carousel generates 4 dark slides (slide2, slide4, slide6, slide8).
The prompts are pipe-separated in column 13 — one segment per dark slide.

Each prompt follows the approved brand template:
  - Deep navy floating panel (#16233A)
  - Specific heading + body text rendered verbatim on the panel
  - Soft backlit indigo-blue glow emanating from BEHIND the panel
  - Wordmark: small light-weight (DR) + large extra-bold LIEW
  - Tagline: ORTHOPAEDICS 360 in brand blue
  - URL: drchienwenliew.com.au
  - Background scene varies by slide topic
"""

import sys
from pathlib import Path
import openpyxl

# Import the carousel content definitions
sys.path.insert(0, str(Path(__file__).parent))
from carousel_content import CAROUSEL_CONTENT

SCRIPT_DIR = Path(__file__).parent
XLSX       = SCRIPT_DIR / "posts_schedule.xlsx"

# Column indices (1-based, matching generate_images_kie.py)
COL_ID      = 1
COL_DATE    = 2
COL_PLATFORM = 4
COL_FORMAT  = 5
COL_KIE_DESC = 13  # "Kie.ai Image Description"

# ── Brand panel template fragments ───────────────────────────────────────────

PANEL_BASE = (
    "Photorealistic editorial social media image. "
    "A deep navy (#16233A) rounded-corner panel floats in the lower 45% of the frame, "
    "inset from all edges — it does not touch the sides or bottom. "
    "A soft diffused radiant glow of deep indigo-blue light emanates from BEHIND the panel "
    "and spreads outward beyond all four panel edges — like the panel is gently backlit, "
    "hovering above the photograph. "
    "The panel is translucent at its very edges but solid dark navy at centre. "
)

WORDMARK = (
    "In the upper-left of the panel: small light-weight '(DR)' immediately followed by "
    "'LIEW' in significantly larger extra-bold white — the size contrast is intentional and prominent. "
    "Directly below: 'ORTHOPAEDICS 360' in small Ship Cove blue (#6D8CBA) tracking-wide caps. "
)

URL_LINE = (
    "At the very bottom of the panel: 'drchienwenliew.com.au' in small white text. "
)

FONT_NOTE = (
    "All panel text uses Neue Montreal typeface — Light, Regular and Bold weights as appropriate. "
    "Text is crisp, clean, perfectly legible. "
)

NO_PEOPLE = (
    "No people, no medical staff, no clinical scenes. "
)

def panel_prompt(heading: str, body: str, background: str, aspect: str = "1:1") -> str:
    """Build a complete Kie.ai prompt for one dark carousel slide."""

    heading_clean = heading.strip().rstrip(".")
    body_clean    = body.strip()

    text_block = (
        f"On the panel, render this text exactly as written — no truncation, no paraphrasing: "
        f"Heading in large bold white: \"{heading_clean}\" "
        f"Body text below in smaller regular white: \"{body_clean}\" "
        f"Render EVERY word completely. Text wraps naturally across lines. "
        f"Absolutely no truncation, no ellipsis, no cutting off. "
    )

    return (
        PANEL_BASE
        + WORDMARK
        + text_block
        + URL_LINE
        + FONT_NOTE
        + f"Background: {background} "
        + NO_PEOPLE
        + f"Aspect ratio {aspect}. "
        "Luxury editorial aesthetic. High-end photography quality."
    )


# ── Background scenes by slide theme ─────────────────────────────────────────

def background_for_slide(heading: str, topic: str, slide_index: int) -> str:
    """Choose a background scene for a slide based on content and position."""
    heading_lower = heading.lower()
    topic_lower   = topic.lower()

    # Surgery / procedure related
    if any(w in heading_lower for w in ["theatre", "surgery", "anaes", "recovery room"]):
        return (
            "warm soft-focus hospital corridor interior, abstract bokeh light, "
            "neutral tones, no people visible"
        )

    # Walking / rehabilitation / mobility
    if any(w in heading_lower for w in ["walk", "physio", "mobility", "independent", "active", "return"]):
        return (
            "sunlit outdoor lifestyle scene — dappled shade, lush greenery, "
            "warm morning light, no people"
        )

    # Pain / symptoms
    if any(w in heading_lower for w in ["pain", "wake", "night", "medication", "waiting"]):
        return (
            "moody twilight landscape — dark blue-grey sky, soft atmospheric mist, "
            "lone park bench in distance, no people"
        )

    # Activities / sport / travel
    if any(w in heading_lower for w in ["golf", "swim", "cycl", "travel", "sport", "avoid"]):
        return (
            "bright lifestyle aerial shot of a coastal walkway or golf course green, "
            "vivid natural colours, no people"
        )

    # Family / relationships
    if any(w in heading_lower for w in ["family", "concern", "help", "support", "independ"]):
        return (
            "warm domestic interior — soft lamp light, cosy lounge, "
            "blurred lifestyle background, no people visible"
        )

    # Longevity / implants / data
    if any(w in heading_lower for w in ["year", "%", "intact", "longevit", "implant"]):
        return (
            "abstract deep blue marble texture with subtle silver veining, "
            "luxury editorial background"
        )

    # Default: vary by position for visual rhythm
    defaults = [
        (
            "warm bokeh photography — shallow depth of field, "
            "soft amber and gold tones, abstract lifestyle"
        ),
        (
            "architectural interior — modern atrium with tall windows, "
            "diffused natural light, no people"
        ),
        (
            "close-up nature texture — smooth river stones, water reflections, "
            "cool blue-green tones"
        ),
        (
            "aerial dawn cityscape with warm horizon glow, "
            "calm atmospheric perspective"
        ),
    ]
    return defaults[slide_index % len(defaults)]


# ── Aspect ratio by platform ──────────────────────────────────────────────────

ASPECT = {
    "instagram": "1:1",
    "facebook":  "1:1",
    "linkedin":  "16:9",
}


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print(f"\n{'='*60}")
    print("  Update Carousel Prompts — writing to posts_schedule.xlsx")
    print(f"{'='*60}\n")

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Posts Schedule"] if "Posts Schedule" in wb.sheetnames else wb.active

    updated = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2):
        cell_id = row[COL_ID - 1]
        if cell_id.value is None:
            continue
        try:
            pid = int(cell_id.value)
        except (ValueError, TypeError):
            continue

        if pid not in CAROUSEL_CONTENT:
            continue

        platform = str(row[COL_PLATFORM - 1].value or "").strip().lower()
        fmt      = str(row[COL_FORMAT - 1].value or "").strip().lower()

        if "carousel" not in fmt and "infographic" not in fmt:
            print(f"  ⚠️  ID {pid}: format '{fmt}' not carousel — skipping")
            skipped += 1
            continue

        content = CAROUSEL_CONTENT[pid]
        topic   = content["topic"]
        slides  = content["slides"]
        aspect  = ASPECT.get(platform, "1:1")

        # Build one prompt per dark slide, pipe-separated
        prompts = []
        for i, slide in enumerate(slides):
            bg = background_for_slide(slide["heading"], topic, i)
            prompt = panel_prompt(
                heading    = slide["heading"],
                body       = slide["body"],
                background = bg,
                aspect     = aspect,
            )
            prompts.append(prompt)

        pipe_joined = " | ".join(prompts)

        # Write to column 13
        row[COL_KIE_DESC - 1].value = pipe_joined

        print(f"  ✅  ID {pid:3d}  {platform:10s}  {topic[:50]}...")
        updated += 1

    wb.save(XLSX)

    print(f"\n{'─'*60}")
    print(f"  Done.  Updated: {updated}  |  Skipped: {skipped}")
    print(f"  Saved to: {XLSX.name}")
    print(f"\n  Next steps:")
    print(f"  1. Run 'Regenerate Posts from Excel.command' to rebuild posts_data.py")
    print(f"  2. Run 'Generate Kie.ai Images.command' (or --date / --from-date flags)")
    print(f"  3. Run 'Push Images to GitHub.command'\n")


if __name__ == "__main__":
    main()
