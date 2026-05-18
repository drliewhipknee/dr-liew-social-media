"""
Regenerate Social Media Post Editor HTML.
Features:
  - Edit Topic, Caption, Hashtags, Website Link, Kie.ai Image Description
  - Auto-detects when a topic changes → orange "Topic Changed" badge
  - "Changed Topics" filter to see only edited posts
  - "Export changed only → JSON" for sending to Claude for content regeneration
  - "Export all → JSON" for full export
  - Reviewed / Mark-as-done checkbox per post
"""
import json, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl

SCRIPT_DIR = Path(__file__).parent
EXCEL_PATH = SCRIPT_DIR / "posts_schedule.xlsx"
OUT_HTML   = SCRIPT_DIR.parent / "Social Media Post Editor.html"

# Read directly from Excel so the HTML is always in sync
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Posts Schedule"]
headers = [cell.value for cell in ws[1]]

raw = []
for row in ws.iter_rows(min_row=2, values_only=False):
    if not any(cell.value for cell in row):
        continue
    d = {}
    for h, cell in zip(headers, row):
        val = cell.value
        d[h] = str(val).strip() if val is not None else ""
    raw.append(d)

def normalise(p):
    return {
        "ID":                         p.get("ID", ""),
        "Date":                       p.get("Date", ""),
        "Day":                        p.get("Day", ""),
        "Platform":                   p.get("Platform", ""),
        "Format":                     p.get("Format", ""),
        "Topic":                      p.get("Topic", ""),
        "Caption":                    p.get("Caption", ""),
        "Hashtags":                   p.get("Hashtags", ""),
        "Website Link":               p.get("Website Link", ""),
        "Image File":                 p.get("Image File", ""),
        "Image Prompt":               p.get("Image Prompt", ""),
        "AI Image Prompt":            p.get("AI Image Prompt", ""),
        "Kie.ai Image Description":   p.get("Kie.ai Image Description", ""),
    }

posts    = [normalise(p) for p in raw]
posts_js = json.dumps(posts, ensure_ascii=False)

total    = len(posts)
li_count = sum(1 for p in posts if p["Platform"] == "LinkedIn")
ig_count = sum(1 for p in posts if p["Platform"] == "Instagram")
fb_count = sum(1 for p in posts if p["Platform"] == "Facebook")

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Dr Liew — Social Media Post Editor</title>
<style>
  :root {{
    --navy: #16233A; --blue: #A2B9D8; --bg: #f4f6f9; --card: #fff;
    --border: #e2e8f0; --text: #1e293b; --muted: #64748b;
    --green: #16a34a; --green-bg: #f0fdf4;
    --orange: #ea580c; --orange-bg: #fff7ed; --orange-border: #fb923c;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
          background: var(--bg); color: var(--text); font-size: 14px; }}

  /* ── Top bar ── */
  .topbar {{ background: var(--navy); color: #fff; padding: 12px 24px;
             display: flex; align-items: center; justify-content: space-between;
             position: sticky; top: 0; z-index: 100; gap: 12px; flex-wrap: wrap; }}
  .topbar-left {{ display: flex; align-items: center; gap: 16px; }}
  .logo {{ font-size: 11px; letter-spacing: .12em; color: var(--blue);
           text-transform: uppercase; font-weight: 500; }}
  .logo span {{ color: #fff; font-size: 18px; font-weight: 700;
                letter-spacing: 0; margin-left: 4px; }}
  .title {{ font-size: 13px; color: rgba(255,255,255,.6); }}
  .topbar-right {{ display: flex; gap: 8px; flex-wrap: wrap; }}
  .export-btn {{ background: var(--blue); color: var(--navy); border: none;
                 padding: 7px 16px; border-radius: 6px; font-size: 13px;
                 font-weight: 600; cursor: pointer; white-space: nowrap; }}
  .export-btn:hover {{ background: #c8d9ec; }}
  .export-changed-btn {{ background: var(--orange); color: #fff; border: none;
                         padding: 7px 16px; border-radius: 6px; font-size: 13px;
                         font-weight: 600; cursor: pointer; white-space: nowrap; }}
  .export-changed-btn:hover {{ background: #c2410c; }}
  .export-changed-btn:disabled {{ background: #d1d5db; color: #9ca3af; cursor: not-allowed; }}

  /* ── Controls bar ── */
  .controls {{ background: #fff; border-bottom: 1px solid var(--border);
               padding: 10px 24px; display: flex; align-items: center;
               gap: 10px; flex-wrap: wrap; position: sticky; top: 53px; z-index: 99; }}
  .filter-btn {{ padding: 5px 14px; border-radius: 20px; border: 1px solid var(--border);
                 background: #fff; cursor: pointer; font-size: 13px; color: var(--muted); }}
  .filter-btn.active {{ background: var(--navy); color: #fff; border-color: var(--navy); }}
  .filter-btn.li.active  {{ background: #0A66C2; border-color: #0A66C2; }}
  .filter-btn.ig.active  {{ background: #C13584; border-color: #C13584; }}
  .filter-btn.fb.active  {{ background: #1877F2; border-color: #1877F2; }}
  .filter-btn.chg.active {{ background: var(--orange); border-color: var(--orange); }}
  .search {{ flex: 1; min-width: 200px; padding: 6px 12px;
             border: 1px solid var(--border); border-radius: 6px;
             font-size: 13px; outline: none; }}
  .search:focus {{ border-color: var(--blue); }}
  .count {{ margin-left: auto; font-size: 12px; color: var(--muted); white-space: nowrap; }}

  /* ── Stats ── */
  .stats {{ display: flex; gap: 10px; padding: 14px 24px; flex-wrap: wrap; }}
  .stat {{ background: #fff; border: 1px solid var(--border); border-radius: 8px;
           padding: 10px 16px; text-align: center; min-width: 90px; }}
  .stat.changed-stat {{ border-color: var(--orange-border); background: var(--orange-bg); }}
  .stat-n {{ font-size: 22px; font-weight: 700; color: var(--navy); }}
  .stat.changed-stat .stat-n {{ color: var(--orange); }}
  .stat-l {{ font-size: 11px; color: var(--muted); margin-top: 2px;
             text-transform: uppercase; letter-spacing: .08em; }}

  /* ── Posts list ── */
  .posts {{ padding: 0 24px 40px; display: flex; flex-direction: column; gap: 10px; }}

  /* ── Card ── */
  .card {{ background: var(--card); border: 1px solid var(--border);
           border-radius: 10px; overflow: hidden; }}
  .card.reviewed     {{ border-left: 3px solid var(--green); }}
  .card.topic-changed {{ border-left: 3px solid var(--orange); }}
  .card.reviewed.topic-changed {{ border-left: 3px solid var(--orange); }}

  .card-header {{ padding: 11px 16px; display: flex; align-items: center; gap: 8px;
                  border-bottom: 1px solid var(--border); background: #fafbfc;
                  cursor: pointer; user-select: none; }}
  .card-header:hover {{ background: #f1f5f9; }}
  .card.topic-changed .card-header {{ background: var(--orange-bg); }}

  .date-badge  {{ font-size: 12px; font-weight: 600; color: var(--muted); min-width: 100px; }}
  .plat-badge  {{ padding: 3px 10px; border-radius: 12px; font-size: 11px; font-weight: 600; color: #fff; }}
  .fmt-badge   {{ padding: 3px 10px; border-radius: 12px; font-size: 11px; font-weight: 500;
                  background: #f1f5f9; color: var(--muted); border: 1px solid var(--border); }}
  .topic-preview {{ flex: 1; font-size: 13px; font-weight: 500; color: var(--text);
                    white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
  .reviewed-badge {{ font-size: 11px; color: var(--green); font-weight: 600;
                     background: var(--green-bg); padding: 2px 8px; border-radius: 10px;
                     white-space: nowrap; }}
  .changed-badge  {{ font-size: 11px; color: var(--orange); font-weight: 600;
                     background: var(--orange-bg); padding: 2px 8px; border-radius: 10px;
                     border: 1px solid var(--orange-border); white-space: nowrap; }}
  .chevron {{ font-size: 16px; color: var(--muted); transition: transform .2s; flex-shrink: 0; }}
  .card.open .chevron {{ transform: rotate(180deg); }}

  /* ── Card body ── */
  .card-body {{ display: none; padding: 16px; }}
  .card.open .card-body {{ display: block; }}

  .field {{ margin-bottom: 14px; }}
  .field label {{ display: block; font-size: 11px; font-weight: 600; color: var(--muted);
                  text-transform: uppercase; letter-spacing: .08em; margin-bottom: 5px; }}
  .field input[type=text] {{ width: 100%; padding: 7px 10px; border: 1px solid var(--border);
                              border-radius: 6px; font-size: 13px; font-family: inherit; outline: none; }}
  .field input[type=text]:focus {{ border-color: var(--blue); }}
  .field input.topic-input {{ font-weight: 600; }}
  .field input.topic-input.changed {{ border-color: var(--orange); background: var(--orange-bg); }}
  .field textarea {{ width: 100%; padding: 8px 10px; border: 1px solid var(--border);
                     border-radius: 6px; font-size: 13px; font-family: inherit;
                     resize: vertical; outline: none; line-height: 1.5; }}
  .field textarea:focus {{ border-color: var(--blue); }}
  .field textarea.caption  {{ min-height: 130px; }}
  .field textarea.hashtags {{ min-height: 60px; font-size: 12px; color: #0A66C2; }}
  .field textarea.img-desc {{ min-height: 80px; background: #f0f8ff; }}

  .topic-hint {{ font-size: 11px; color: var(--orange); margin-top: 4px; display: none; }}
  .topic-hint.visible {{ display: block; }}

  .slides-container {{ border: 1px solid var(--border); border-radius: 8px; overflow: hidden; }}
  .slide-item {{ border-bottom: 1px solid var(--border); padding: 10px 12px; }}
  .slide-item:last-child {{ border-bottom: none; }}
  .slide-label {{ font-size: 11px; font-weight: 600; color: var(--muted);
                  text-transform: uppercase; letter-spacing: .08em; margin-bottom: 5px; }}
  .slide-label span {{ background: var(--navy); color: #fff; padding: 1px 7px;
                       border-radius: 10px; font-size: 10px; margin-right: 4px; }}
  .slide-item textarea {{ width: 100%; border: none; resize: vertical; font-size: 13px;
                          font-family: inherit; outline: none; background: transparent;
                          min-height: 60px; line-height: 1.5; }}

  .card-footer {{ display: flex; align-items: center; justify-content: space-between;
                  padding: 10px 16px; border-top: 1px solid var(--border); background: #fafbfc; }}
  .review-label {{ display: flex; align-items: center; gap: 6px; font-size: 13px;
                   color: var(--muted); cursor: pointer; }}
  .review-label input {{ cursor: pointer; }}
  .post-id {{ font-size: 11px; color: #cbd5e1; }}

  /* ── Changed-topic notice banner ── */
  .changed-banner {{ background: var(--orange-bg); border: 1px solid var(--orange-border);
                     border-radius: 8px; padding: 10px 14px; margin-bottom: 14px;
                     font-size: 13px; color: var(--orange); display: none; }}
  .changed-banner.visible {{ display: block; }}
</style>
</head>
<body>

<div class="topbar">
  <div class="topbar-left">
    <div class="logo">(DR)<span>LIEW</span></div>
    <div class="title">Social Media Post Editor</div>
  </div>
  <div class="topbar-right">
    <button class="export-changed-btn" id="exportChangedBtn"
            onclick="exportChangedJSON()" disabled>
      Export changed topics → JSON
    </button>
    <button class="export-btn" onclick="exportAllJSON()">Export all → JSON</button>
  </div>
</div>

<div class="controls">
  <button class="filter-btn active" data-filter="all"       onclick="setFilter('all',this)">All posts</button>
  <button class="filter-btn li"     data-filter="LinkedIn"  onclick="setFilter('LinkedIn',this)">LinkedIn</button>
  <button class="filter-btn ig"     data-filter="Instagram" onclick="setFilter('Instagram',this)">Instagram</button>
  <button class="filter-btn fb"     data-filter="Facebook"  onclick="setFilter('Facebook',this)">Facebook</button>
  <button class="filter-btn chg"    data-filter="changed"   onclick="setFilter('changed',this)" id="changedFilterBtn">
    Changed topics (0)
  </button>
  <input class="search" type="text" placeholder="Search topic, caption, date…"
         id="search" oninput="applyFilters()">
  <div class="count" id="countDisplay"></div>
</div>

<div class="stats">
  <div class="stat"><div class="stat-n">{total}</div><div class="stat-l">Total Posts</div></div>
  <div class="stat"><div class="stat-n">{li_count}</div><div class="stat-l">LinkedIn</div></div>
  <div class="stat"><div class="stat-n">{ig_count}</div><div class="stat-l">Instagram</div></div>
  <div class="stat"><div class="stat-n">{fb_count}</div><div class="stat-l">Facebook</div></div>
  <div class="stat"><div class="stat-n" id="s-reviewed">0</div><div class="stat-l">Reviewed</div></div>
  <div class="stat changed-stat">
    <div class="stat-n" id="s-changed">0</div>
    <div class="stat-l">Topics changed</div>
  </div>
</div>

<div class="posts" id="postsContainer"></div>

<script>
const POSTS = {posts_js};

// Store original topics so we can detect changes
const origTopics = {{}};
POSTS.forEach(function(p) {{ origTopics[p.ID] = p.Topic || ''; }});

// State: working copy of all posts
const state = POSTS.map(function(p) {{
  return Object.assign({{}}, p, {{_reviewed: false, _topicChanged: false}});
}});

const PLAT_COLORS = {{LinkedIn:'#0A66C2', Instagram:'#C13584', Facebook:'#1877F2'}};
let currentFilter = 'all', currentSearch = '';

// ── DOM helpers ──────────────────────────────────────
function mk(tag, cls, attrs) {{
  var el = document.createElement(tag);
  if (cls) el.className = cls;
  if (attrs) Object.keys(attrs).forEach(function(k) {{ el.setAttribute(k, attrs[k]); }});
  return el;
}}

// ── Changed-topic tracking ───────────────────────────
function setTopicChanged(idx, changed) {{
  state[idx]._topicChanged = changed;
  var card = document.getElementById('card-' + idx);

  // Orange left border
  card.classList.toggle('topic-changed', changed);

  // Badge in header
  var hdr  = card.querySelector('.card-header');
  var badge = hdr.querySelector('.changed-badge');
  if (changed && !badge) {{
    badge = mk('div', 'changed-badge');
    badge.textContent = '✎ Topic changed';
    hdr.querySelector('.chevron').before(badge);
  }} else if (!changed && badge) {{
    badge.remove();
  }}

  // Highlight the topic input itself
  var inp = card.querySelector('.topic-input');
  if (inp) inp.classList.toggle('changed', changed);

  // Hint text under topic field
  var hint = card.querySelector('.topic-hint');
  if (hint) hint.classList.toggle('visible', changed);

  updateChangedCount();
}}

function updateChangedCount() {{
  var n = state.filter(function(p) {{ return p._topicChanged; }}).length;
  document.getElementById('s-changed').textContent = n;
  document.getElementById('changedFilterBtn').textContent =
    'Changed topics (' + n + ')';
  var btn = document.getElementById('exportChangedBtn');
  btn.disabled = n === 0;
  btn.textContent = n > 0
    ? 'Export ' + n + ' changed topic' + (n === 1 ? '' : 's') + ' → JSON'
    : 'Export changed topics → JSON';
  applyFilters();
}}

// ── Card renderer ────────────────────────────────────
function renderCard(idx) {{
  var p = state[idx];
  var kd = p['Kie.ai Image Description'] || '';
  var fmt = (p.Format || '').toLowerCase();
  var isCarousel = fmt === 'carousel' || fmt === 'infographic';
  var platColor  = PLAT_COLORS[p.Platform] || '#64748b';
  var hashCount  = p.Hashtags ? p.Hashtags.trim().split(/\\s+/).length : 0;

  var card = mk('div', 'card' + (p._reviewed ? ' reviewed' : ''));
  card.id = 'card-' + idx;

  // ── Header ──
  var hdr = mk('div', 'card-header');
  hdr.onclick = function() {{ toggleCard(idx); }};

  var dateBadge = mk('div', 'date-badge');
  dateBadge.textContent = (p.Date || '') + (p.Day ? ' (' + p.Day.slice(0,3) + ')' : '');

  var platBadge = mk('div', 'plat-badge');
  platBadge.style.background = platColor;
  platBadge.textContent = p.Platform;

  var fmtBadge = mk('div', 'fmt-badge');
  fmtBadge.textContent = p.Format;

  var topicPreview = mk('div', 'topic-preview');
  topicPreview.id = 'tp-' + idx;
  topicPreview.textContent = p.Topic || '';

  var chevron = mk('div', 'chevron');
  chevron.textContent = '▾';

  hdr.appendChild(dateBadge);
  hdr.appendChild(platBadge);
  hdr.appendChild(fmtBadge);
  hdr.appendChild(topicPreview);
  if (p._reviewed) {{
    var rb = mk('div', 'reviewed-badge');
    rb.textContent = '✓ Reviewed';
    hdr.appendChild(rb);
  }}
  hdr.appendChild(chevron);

  // ── Body ──
  var body = mk('div', 'card-body');

  // Changed-topic banner
  var banner = mk('div', 'changed-banner');
  banner.textContent = '✎ Topic changed — export this post and send to Claude to regenerate the caption, hashtags and image description.';
  body.appendChild(banner);

  // Topic field
  body.appendChild(makeField('Topic', function() {{
    var inp = mk('input', 'topic-input', {{type:'text'}});
    inp.value = p.Topic || '';
    inp.addEventListener('input', function() {{
      state[idx].Topic = this.value;
      document.getElementById('tp-' + idx).textContent = this.value;
      var changed = this.value.trim() !== origTopics[p.ID].trim();
      setTopicChanged(idx, changed);
      // Show/hide banner
      banner.classList.toggle('visible', changed);
    }});
    // Hint
    var hint = mk('div', 'topic-hint');
    hint.textContent = '⚠ Topic changed from original. Export and send to Claude to regenerate content.';
    var wrap = mk('div');
    wrap.appendChild(inp);
    wrap.appendChild(hint);
    return wrap;
  }}));

  // Caption field
  body.appendChild(makeField('Caption', function() {{
    var ta = mk('textarea', 'caption');
    ta.value = p.Caption || '';
    ta.addEventListener('input', function() {{ state[idx].Caption = this.value; }});
    return ta;
  }}));

  // Hashtags field
  body.appendChild(makeField('Hashtags', function() {{
    var ta = mk('textarea', 'hashtags');
    ta.value = p.Hashtags || '';
    ta.addEventListener('input', function() {{ state[idx].Hashtags = this.value; }});
    return ta;
  }}));

  // Website Link field
  body.appendChild(makeField('Website Link', function() {{
    var inp = mk('input', null, {{type:'text'}});
    inp.value = p['Website Link'] || '';
    inp.addEventListener('input', function() {{ state[idx]['Website Link'] = this.value; }});
    return inp;
  }}));

  // Kie.ai Image Description field
  var imgLbl = isCarousel
    ? 'Kie.ai Image Description (carousel — slides separated by |)'
    : 'Kie.ai Image Description';
  body.appendChild(makeField(imgLbl, function() {{
    if (isCarousel) {{
      return buildSlidesWidget(idx, kd);
    }} else {{
      var ta = mk('textarea', 'img-desc');
      ta.value = kd;
      ta.addEventListener('input', function() {{
        state[idx]['Kie.ai Image Description'] = this.value;
      }});
      return ta;
    }}
  }}));

  // ── Footer ──
  var footer = mk('div', 'card-footer');

  var reviewLabel = mk('label', 'review-label');
  var cb = mk('input', null, {{type:'checkbox'}});
  if (p._reviewed) cb.checked = true;
  cb.addEventListener('change', function() {{ toggleReview(idx, this.checked); }});
  reviewLabel.appendChild(cb);
  reviewLabel.appendChild(document.createTextNode(' Mark as reviewed'));

  var postId = mk('div', 'post-id');
  postId.textContent = 'ID: ' + p.ID + '  |  ' + hashCount + ' hashtags';

  footer.appendChild(reviewLabel);
  footer.appendChild(postId);
  body.appendChild(footer);

  card.appendChild(hdr);
  card.appendChild(body);
  return card;
}}

function makeField(labelText, buildInput) {{
  var wrap = mk('div', 'field');
  var lbl  = mk('label');
  lbl.textContent = labelText;
  wrap.appendChild(lbl);
  var input = buildInput();
  wrap.appendChild(input);
  return wrap;
}}

function buildSlidesWidget(idx, desc) {{
  var sep    = desc.includes('|SLIDE|') ? '|SLIDE|' : ' | ';
  var slides = desc.split(sep);
  var container = mk('div', 'slides-container');
  slides.forEach(function(s, i) {{
    var item  = mk('div', 'slide-item');
    var lbl   = mk('div', 'slide-label');
    var badge = mk('span');
    badge.textContent = i + 1;
    lbl.appendChild(badge);
    lbl.appendChild(document.createTextNode(' Slide ' + (i + 1)));
    var ta = mk('textarea');
    ta.value = s.trim();
    ta.addEventListener('input', function() {{ updateSlide(idx, i, this.value, sep); }});
    item.appendChild(lbl);
    item.appendChild(ta);
    container.appendChild(item);
  }});
  return container;
}}

function updateSlide(idx, slideIdx, value, sep) {{
  var slides = (state[idx]['Kie.ai Image Description'] || '').split(sep);
  slides[slideIdx] = value;
  state[idx]['Kie.ai Image Description'] = slides.join(sep);
}}

// ── Card open/close ──────────────────────────────────
function toggleCard(idx) {{
  document.getElementById('card-' + idx).classList.toggle('open');
}}

// ── Review toggle ─────────────────────────────────────
function toggleReview(idx, checked) {{
  state[idx]._reviewed = checked;
  var card = document.getElementById('card-' + idx);
  card.classList.toggle('reviewed', checked);
  var hdr   = card.querySelector('.card-header');
  var badge = hdr.querySelector('.reviewed-badge');
  if (checked && !badge) {{
    badge = mk('div', 'reviewed-badge');
    badge.textContent = '✓ Reviewed';
    hdr.querySelector('.chevron').before(badge);
  }} else if (!checked && badge) {{
    badge.remove();
  }}
  document.getElementById('s-reviewed').textContent =
    state.filter(function(p) {{ return p._reviewed; }}).length;
}}

// ── Filters ──────────────────────────────────────────
function setFilter(val, btn) {{
  currentFilter = val;
  document.querySelectorAll('.filter-btn').forEach(function(b) {{
    b.classList.remove('active');
  }});
  btn.classList.add('active');
  applyFilters();
}}

function applyFilters() {{
  currentSearch = document.getElementById('search').value.toLowerCase();
  var visible = 0;
  state.forEach(function(p, idx) {{
    var card = document.getElementById('card-' + idx);
    if (!card) return;
    var matchPlat    = currentFilter === 'all' || p.Platform === currentFilter;
    var matchChanged = currentFilter !== 'changed' || p._topicChanged;
    var matchSearch  = !currentSearch
      || (p.Topic   || '').toLowerCase().includes(currentSearch)
      || (p.Caption || '').toLowerCase().includes(currentSearch)
      || (p.Date    || '').includes(currentSearch)
      || (p.Format  || '').toLowerCase().includes(currentSearch);
    var show = matchPlat && matchChanged && matchSearch;
    card.style.display = show ? '' : 'none';
    if (show) visible++;
  }});
  document.getElementById('countDisplay').textContent =
    visible + ' of ' + state.length + ' posts';
}}

// ── Export functions ──────────────────────────────────
function buildExportRow(p) {{
  return {{
    'ID':                         p.ID,
    'Topic':                      p.Topic,
    'Caption':                    p.Caption,
    'Hashtags':                   p.Hashtags,
    'Website Link':               p['Website Link'],
    'Kie.ai Image Description':   p['Kie.ai Image Description'],
    '_topicChanged':              p._topicChanged || false
  }};
}}

function downloadJSON(data, filename) {{
  var blob = new Blob([JSON.stringify(data, null, 2)], {{type:'application/json'}});
  var a    = document.createElement('a');
  a.href   = URL.createObjectURL(blob);
  a.download = filename;
  a.click();
}}

function exportAllJSON() {{
  downloadJSON(state.map(buildExportRow), 'posts_edited.json');
}}

function exportChangedJSON() {{
  var changed = state.filter(function(p) {{ return p._topicChanged; }});
  if (!changed.length) {{ alert('No topics have been changed yet.'); return; }}
  downloadJSON(changed.map(buildExportRow), 'posts_changed_topics.json');
}}

// ── Init ─────────────────────────────────────────────
(function() {{
  var container = document.getElementById('postsContainer');
  state.forEach(function(_, idx) {{ container.appendChild(renderCard(idx)); }});
  applyFilters();
}})();
</script>
</body>
</html>"""

OUT_HTML.write_text(html, encoding="utf-8")
print(f"Generated: {OUT_HTML}")
print(f"Posts: {total} ({li_count} LI, {ig_count} IG, {fb_count} FB)")
