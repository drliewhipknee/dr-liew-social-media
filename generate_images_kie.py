#!/usr/bin/env python3
"""
generate_images_kie.py
Generates complete branded post images via Kie.ai GPT Image 2.

Reads the "Kie.ai Image Description" column from posts_schedule.xlsx,
sends each description to Kie.ai gpt-image-2-text-to-image, and saves
the finished image to generated_images/.

For carousel posts the description is pipe-separated (one prompt per
dark slide). Each segment generates one slide image.

Usage:
    python3 generate_images_kie.py                                          # all pending posts
    python3 generate_images_kie.py --date 2026-05-05                        # single date
    python3 generate_images_kie.py --id 85                                  # single post ID
    python3 generate_images_kie.py --from-date 2026-05-01                   # from date onwards
    python3 generate_images_kie.py --from-date 2026-05-01 --to-date 2026-05-31  # one month only
    python3 generate_images_kie.py --force                                  # regenerate even if done
"""

import argparse, json, logging, sys, time
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

try:
    from PIL import Image, ImageDraw
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "Pillow", "--break-system-packages", "-q"])
    from PIL import Image, ImageDraw

# ── Config ────────────────────────────────────────────────────────────────────
KIE_API_KEY  = "961df6a53af50676575fee66b918b68c"
KIE_BASE     = "https://api.kie.ai/api/v1"
MODEL        = "gpt-image-2-text-to-image"
POLL_INTERVAL = 6    # seconds between status checks
MAX_WAIT      = 180  # seconds before giving up on a task

SCRIPT_DIR    = Path(__file__).parent
XLSX          = SCRIPT_DIR / "posts_schedule.xlsx"
OUT_DIR       = SCRIPT_DIR / "generated_images"
PROGRESS_FILE = OUT_DIR / "_kie_progress.json"
OUT_DIR.mkdir(exist_ok=True)

ACTIVE_FROM_ID = 79

# Headshot for Facebook circle compositing
HEADSHOT_PATH = SCRIPT_DIR.parent / "Liew headshot.jpg"


def composite_headshot(image_path: Path) -> None:
    """
    Composite Dr Liew's headshot into the reserved bottom-right circle
    on a Facebook image. The circle occupies ~21.5% of image width,
    positioned ~10.7% from the right and bottom edges.
    Saves back to the same path.
    """
    if not HEADSHOT_PATH.exists():
        log.warning(f"  Headshot not found at {HEADSHOT_PATH} — skipping composite")
        return
    try:
        base = Image.open(image_path).convert("RGBA")
        W, H = base.size

        # Circle dimensions — proportional to image size
        diameter = round(W * 0.215)          # ~220px on 1024px image
        margin   = round(W * 0.107)          # ~110px from edges
        cx = W - margin - diameter // 2      # circle centre x
        cy = H - margin - diameter // 2      # circle centre y
        box = (cx - diameter // 2, cy - diameter // 2,
               cx + diameter // 2, cy + diameter // 2)

        # Prepare headshot: centre-crop to square, resize to circle diameter
        hs = Image.open(HEADSHOT_PATH).convert("RGBA")
        hs_w, hs_h = hs.size
        crop_side = min(hs_w, hs_h)
        left  = (hs_w - crop_side) // 2
        top   = (hs_h - crop_side) // 2
        hs = hs.crop((left, top, left + crop_side, top + crop_side))
        hs = hs.resize((diameter, diameter), Image.LANCZOS)

        # Apply circular mask to headshot
        mask = Image.new("L", (diameter, diameter), 0)
        ImageDraw.Draw(mask).ellipse((0, 0, diameter - 1, diameter - 1), fill=255)
        hs.putalpha(mask)

        # Draw a 3px light-blue border ring over the circle area
        ring = Image.new("RGBA", (diameter + 6, diameter + 6), (0, 0, 0, 0))
        ImageDraw.Draw(ring).ellipse(
            (0, 0, diameter + 5, diameter + 5),
            outline=(162, 185, 216, 220), width=3
        )

        # Paste headshot then ring onto base image
        base.paste(hs, (box[0], box[1]), hs)
        base.paste(ring, (box[0] - 3, box[1] - 3), ring)

        base.convert("RGB").save(image_path, "JPEG", quality=95)
        log.info(f"  📸  Headshot composited into {image_path.name}")
    except Exception as e:
        log.warning(f"  Headshot composite failed: {e}")


ASPECT = {
    "instagram": "1:1",
    "linkedin":  "16:9",
    "facebook":  "1:1",
}

# Column indices (1-based)
COL = {
    "ID": 1, "Date": 2, "Day": 3, "Platform": 4, "Format": 5, "Topic": 6,
    "Caption": 7, "Hashtags": 8, "Website Link": 9,
    "Image File": 10, "Image Prompt": 11, "AI Image Prompt": 12,
    "Kie.ai Image Description": 13,
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

HEADERS = {
    "Authorization": f"Bearer {KIE_API_KEY}",
    "Content-Type":  "application/json",
}


# ── Progress tracking ─────────────────────────────────────────────────────────
def load_progress():
    if PROGRESS_FILE.exists():
        return json.loads(PROGRESS_FILE.read_text())
    return {}

def save_progress(p):
    PROGRESS_FILE.write_text(json.dumps(p, indent=2))


# ── Kie.ai API ────────────────────────────────────────────────────────────────
def create_task(prompt: str, aspect_ratio: str) -> str | None:
    """Submit a generation task. Returns taskId or None on failure."""
    try:
        resp = requests.post(
            f"{KIE_BASE}/jobs/createTask",
            headers=HEADERS,
            json={
                "model":       MODEL,
                "callBackUrl": "",
                "input": {
                    "prompt":       prompt,
                    "aspect_ratio": aspect_ratio,
                },
            },
            timeout=30,
        )
        resp.raise_for_status()
        body = resp.json()

        # Navigate common response shapes
        data = body.get("data") or body
        task_id = (
            data.get("taskId")
            or data.get("task_id")
            or data.get("id")
            or body.get("taskId")
        )
        if not task_id:
            log.warning(f"  Unexpected createTask response: {body}")
        return task_id
    except Exception as e:
        log.error(f"  createTask failed: {e}")
        return None


def poll_task(task_id: str) -> str | None:
    """Poll until done. Returns image URL or None."""
    deadline = time.time() + MAX_WAIT
    while time.time() < deadline:
        time.sleep(POLL_INTERVAL)
        try:
            resp = requests.get(
                f"{KIE_BASE}/jobs/recordInfo",
                headers=HEADERS,
                params={"taskId": task_id},
                timeout=30,
            )
            resp.raise_for_status()
            body = resp.json()
            data = body.get("data") or body

            state = (
                data.get("state")
                or data.get("status")
                or data.get("taskStatus")
                or ""
            ).upper()

            log.info(f"    status: {state}")

            if state in ("SUCCESS", "COMPLETED", "DONE", "FINISH"):
                # Extract URL from resultJson or direct fields
                result_json_str = data.get("resultJson") or "{}"
                if isinstance(result_json_str, str):
                    result_json = json.loads(result_json_str)
                else:
                    result_json = result_json_str

                urls = (
                    result_json.get("resultUrls")
                    or result_json.get("images")
                    or data.get("resultUrls")
                    or data.get("images")
                    or []
                )
                if isinstance(urls, list) and urls:
                    url = urls[0]
                    return url.get("url") if isinstance(url, dict) else url
                # Try direct URL fields
                return data.get("url") or result_json.get("url")

            elif state in ("FAILED", "FAIL", "ERROR", "CANCELLED"):
                # Log full response so we can diagnose API rejections
                fail_msg = (
                    data.get("failReason")
                    or data.get("error")
                    or data.get("message")
                    or data.get("msg")
                    or body.get("message")
                    or body.get("msg")
                    or ""
                )
                log.warning(f"  Task {task_id} failed ({state}): {fail_msg or 'no detail'}")
                log.debug(f"  Full fail body: {body}")
                return None

        except Exception as e:
            log.warning(f"  Poll error: {e}")

    log.warning(f"  Task {task_id} timed out after {MAX_WAIT}s")
    return None


def download(url: str, out_path: Path) -> bool:
    """Download image to out_path. Returns True on success."""
    try:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        out_path.write_bytes(resp.content)
        log.info(f"  ✅  Saved {out_path.name} ({len(resp.content)//1024} KB)")
        return True
    except Exception as e:
        log.error(f"  Download failed: {e}")
        return False


def generate_one(prompt: str, aspect_ratio: str, out_path: Path) -> bool:
    """Full flow: create → poll → download. Returns True on success."""
    log.info(f"  → Creating task ({aspect_ratio}) ...")
    task_id = create_task(prompt, aspect_ratio)
    if not task_id:
        return False
    log.info(f"  → Task ID: {task_id}. Polling ...")
    url = poll_task(task_id)
    if not url:
        return False
    return download(url, out_path)


# ── Excel loading ─────────────────────────────────────────────────────────────
def load_posts(target_date=None, target_id=None, from_date=None, to_date=None, target_platform=None):
    wb = openpyxl.load_workbook(XLSX)
    # Try "Posts Schedule" sheet first, fall back to active
    ws = wb["Posts Schedule"] if "Posts Schedule" in wb.sheetnames else wb.active

    posts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[COL["ID"] - 1]
        if pid is None:
            continue
        try:
            pid = int(pid)
        except (ValueError, TypeError):
            continue
        if pid < ACTIVE_FROM_ID:
            continue

        dt = row[COL["Date"] - 1]
        if isinstance(dt, datetime):
            d_str = dt.strftime("%Y-%m-%d")
        else:
            d_str = str(dt).strip() if dt else ""

        if target_date and d_str != target_date:
            continue
        if target_id and pid != target_id:
            continue
        if from_date and d_str < from_date:
            continue
        if to_date and d_str > to_date:
            continue

        platform_val = str(row[COL["Platform"] - 1] or "").strip().lower()
        if target_platform and target_platform.lower() not in platform_val:
            continue

        posts.append({
            "id":          pid,
            "date":        d_str,
            "platform":    platform_val,
            "format":      str(row[COL["Format"]   - 1] or "").strip().lower(),
            "kie_desc":    str(row[COL["Kie.ai Image Description"] - 1] or "").strip(),
        })

    return posts


# ── Filename helpers ──────────────────────────────────────────────────────────
def single_filename(post):
    return f"{post['date']}-{post['platform']}.jpg"

def slide_filename(post, suffix):
    return f"{post['date']}-{post['platform']}-{suffix}.jpg"


# ── Main generation loop ──────────────────────────────────────────────────────
def run(posts, force=False):
    progress = load_progress()
    aspect   = ASPECT.get

    total = ok = skipped = failed = 0

    for post in posts:
        pid      = post["id"]
        platform = post["platform"]
        fmt      = post["format"]
        desc     = post["kie_desc"]

        if not desc:
            log.warning(f"ID {pid} ({post['date']}): no Kie.ai description — skipping")
            skipped += 1
            continue

        is_carousel = "carousel" in fmt or "infographic" in fmt
        ar = aspect(platform, "1:1")

        if is_carousel:
            # Split pipe-separated descriptions → one image per slide
            segments = [s.strip() for s in desc.split("|") if s.strip()]
            # Carousel dark slides are slide2, slide4, slide6, slide8
            dark_slide_nums = [2, 4, 6, 8]

            for i, segment in enumerate(segments):
                suf = f"slide{dark_slide_nums[i % len(dark_slide_nums)]}"
                fname = slide_filename(post, suf)
                key   = f"{pid}-{suf}"
                out   = OUT_DIR / fname

                if not force and progress.get(key) == "done" and out.exists():
                    log.info(f"ID {pid} {suf}: already done — skip")
                    skipped += 1
                    continue

                log.info(f"ID {pid} ({post['date']}) {platform} {suf}:")
                total += 1
                if generate_one(segment, ar, out):
                    progress[key] = "done"
                    save_progress(progress)
                    ok += 1
                else:
                    progress[key] = "error"
                    save_progress(progress)
                    failed += 1

            # Also generate slide1 (cover) from the first segment + cover styling note
            cover_key = f"{pid}-slide1"
            cover_out = OUT_DIR / slide_filename(post, "slide1")
            if force or progress.get(cover_key) != "done" or not cover_out.exists():
                cover_prompt = segments[0] + " COVER SLIDE — hero image, bold and striking."
                log.info(f"ID {pid} ({post['date']}) {platform} slide1 (cover):")
                total += 1
                if generate_one(cover_prompt, ar, cover_out):
                    progress[cover_key] = "done"
                    save_progress(progress)
                    ok += 1
                else:
                    progress[cover_key] = "error"
                    save_progress(progress)
                    failed += 1

        else:
            # Single image post
            fname = single_filename(post)
            key   = str(pid)
            out   = OUT_DIR / fname

            if not force and progress.get(key) == "done" and out.exists():
                log.info(f"ID {pid} ({post['date']}) {platform}: already done — skip")
                skipped += 1
                continue

            log.info(f"ID {pid} ({post['date']}) {platform} [{fmt}]:")
            total += 1
            if generate_one(desc, ar, out):
                # Composite Dr Liew's headshot into Facebook images
                if platform == "facebook":
                    composite_headshot(out)
                progress[key] = "done"
                save_progress(progress)
                ok += 1
            else:
                progress[key] = "error"
                save_progress(progress)
                failed += 1

    log.info(
        f"\n{'='*50}\n"
        f"Done.  Generated: {ok}  |  Skipped: {skipped}  |  Failed: {failed}  |  Total attempted: {total}\n"
        f"Images saved to: {OUT_DIR}\n"
        f"{'='*50}"
    )


# ── CLI ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Generate post images via Kie.ai GPT Image 2")
    parser.add_argument("--date",      help="Generate for a specific date (YYYY-MM-DD)")
    parser.add_argument("--id",        type=int, help="Generate for a specific post ID")
    parser.add_argument("--from-date", help="Generate from this date onwards (YYYY-MM-DD)")
    parser.add_argument("--to-date",   help="Generate up to and including this date (YYYY-MM-DD)")
    parser.add_argument("--force",     action="store_true", help="Regenerate even if already done")
    parser.add_argument("--platform",  help="Filter by platform (e.g. facebook, linkedin, instagram)")
    args = parser.parse_args()

    posts = load_posts(
        target_date=args.date,
        target_id=args.id,
        from_date=args.from_date,
        to_date=args.to_date,
        target_platform=args.platform,
    )

    if not posts:
        log.info("No matching posts found.")
        return

    log.info(f"Found {len(posts)} posts to process.")
    run(posts, force=args.force)


if __name__ == "__main__":
    main()
