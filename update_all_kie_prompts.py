#!/usr/bin/env python3
"""
update_all_kie_prompts.py

Updates every active post's "Kie.ai Image Description" in posts_schedule.xlsx
with a high-quality, world-class art-directed prompt.

Platform layouts:
  Instagram (1:1)  — Photography top 58%, floating navy panel bottom 42%
  LinkedIn  (16:9) — Floating navy panel left 42%, editorial photography right 58%
  Facebook  (1:1)  — Photography top 58%, floating panel bottom 42%,
                     reserved circular area bottom-right for Dr Liew headshot

Usage:
    python3 update_all_kie_prompts.py
"""

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl',
                    '--break-system-packages', '-q'])
    import openpyxl

XLSX           = Path(__file__).parent / 'posts_schedule.xlsx'
ACTIVE_FROM_ID = 79


# ══════════════════════════════════════════════════════════════════════════════
# BACKGROUND SCENE LIBRARY
# Each category has multiple variants; post_id % len(variants) selects one.
# This ensures variety without being random (deterministic, reproducible).
# ══════════════════════════════════════════════════════════════════════════════

SCENES = {

    'hip': [
        ("Warm private consulting room with a polished titanium hip joint anatomical model on a "
         "timber desk, soft desk lamp casting warm pools of light, architectural shelving lined with "
         "medical texts behind, no people, premium private practice aesthetic."),
        ("Person (back to camera, anonymous, face entirely not visible) walking confidently along a "
         "wide sunlit Adelaide park path bordered by tall eucalyptus trees, golden hour light, long "
         "shadows stretching behind them, warm and hopeful wellness lifestyle feel, no clinical elements."),
        ("Adelaide Hills rolling vineyard landscape at golden hour, warm amber and deep green tones, "
         "rows of vines catching the last light, peaceful and distinctly South Australian, no people."),
        ("Elegant private hospital atrium, soaring glass and timber architecture, warm ambient light "
         "reflecting off polished limestone floors, a single potted ficus in the distance, no people, "
         "premium private healthcare aesthetic."),
    ],

    'knee': [
        ("Empty lap pool at dawn in a luxury aquatic rehabilitation facility, crystal clear blue-green "
         "water, reflected morning light dancing across the pool tiles, premium wellness aesthetic, "
         "no people."),
        ("Person (completely anonymous, back to camera, face not visible, wearing neutral activewear) "
         "descending broad stone garden steps with a confident stride, warm dappled afternoon light "
         "through overhanging trees, no clinical elements, lifestyle and freedom."),
        ("Contemporary private medical suite with warm timber joinery, deep leather chair, architectural "
         "desk lamp casting warm light, abstract artwork on white walls, no people, prestigious and calm."),
        ("Outdoor walking trail through the Adelaide Botanic Gardens at golden hour, dappled amber light "
         "through tall canopy trees, nobody in frame, peaceful and aspirational."),
    ],

    'recovery': [
        ("Bright contemporary home living room, morning light streaming through large white-framed windows "
         "onto a comfortable armchair with a soft linen cushion, a glass of water on a side table, calm "
         "and restorative domestic scene, no people."),
        ("Person (anonymous, back to camera only, face not visible) doing a gentle morning stretch on a "
         "warm timber deck surrounded by lush garden bokeh, soft golden sunrise light, wellness lifestyle "
         "feel, no clinical elements."),
        ("Modern rehabilitation studio, floor-to-ceiling windows with filtered morning light casting long "
         "warm bars across clean timber floors, minimal exercise equipment artfully arranged in the "
         "background, no people, premium and aspirational."),
        ("Sunlit outdoor terrace of a private wellness centre, potted white agapanthus, warm morning light "
         "on clean stone surfaces, restorative and calm, no people."),
    ],

    'physio': [
        ("Anonymous healthcare figure — face entirely not visible, cropped at collarbone, wearing soft "
         "sage-toned clothing — gently supporting a patient's knee with both hands on a warm timber bench "
         "in a sunlit private garden, rich warm amber light streaming through bokeh foliage, no clinical "
         "elements, editorial production quality."),
        ("Warm wellness studio interior, large windows with soft diffused morning light, a practitioner's "
         "hands (anonymous, no face, no identity) guiding a gentle leg extension, clean timber floors, "
         "lush plants in background, premium lifestyle healthcare feel."),
        ("Sun-drenched outdoor terrace, a rehabilitation professional's hands (anonymous, cropped, face "
         "not visible) working on a patient's shoulder, warm afternoon bokeh of garden behind them, "
         "no clinical elements, non-hospital, lifestyle and warmth."),
    ],

    'exercise': [
        ("Person (anonymous, silhouetted, back to camera) cycling along an Adelaide coastal path at "
         "sunrise, warm amber and rose-gold light on calm water, freedom and active lifestyle, no faces."),
        ("Contemporary wellness studio exterior, warm light through glass walls, blurred figures "
         "exercising inside visible from outside, premium lifestyle gym aesthetic, no close faces."),
        ("Empty outdoor stadium steps in early morning mist, long diagonal shadows, golden side light, "
         "no people, disciplined and aspirational feel."),
        ("Outdoor swimming pool lane ropes seen from above, crystal blue-green water, morning light "
         "creating patterns on the bottom, premium aquatic lifestyle, no people."),
    ],

    'wellness': [
        ("Vibrant Adelaide Central Market produce — warm morning light on fresh colourful vegetables "
         "and fruit, no faces, healthy and nourishing lifestyle feel."),
        ("Clean white marble kitchen bench, soft morning light from a window, a tall glass of water "
         "beside fresh cut fruit, simple and calming, no people."),
        ("Person (anonymous, back to camera, face not visible, small in frame) walking slowly along "
         "a Glenelg beach shoreline at sunrise, vast warm amber sky reflected on calm water, sense "
         "of progress and hope."),
        ("Lush home herb garden in terracotta pots on a sunlit windowsill, morning light, healthy "
         "lifestyle and nourishment, no people."),
    ],

    'supplements': [
        ("Clean white marble surface with soft morning window light, a single glass of water beside "
         "a small elegant arrangement of natural supplements in a ceramic bowl, minimal and luxurious, "
         "no people."),
        ("Warm kitchen bench, morning light through sheer white linen curtains, fresh herbs and a "
         "glass of water, soft focus background of a bright kitchen, healthy domestic lifestyle, "
         "no people."),
    ],

    'research': [
        ("Sleek modern glass office building facade at dusk, warm amber interior lights glowing "
         "through floor-to-ceiling windows, Adelaide city reflected in the glass, professional "
         "and aspirational, no people."),
        ("University medical library, tall shelves of bound medical journals, warm reading lamp "
         "casting a pool of amber light on a timber desk, scholarly and prestigious, no people."),
        ("High-resolution data visualisation displayed on a large curved monitor in a contemporary "
         "darkened office, warm side lighting, blue data glow on the desk, analytical and confident, "
         "no people."),
        ("Modern medical conference centre lobby, clean architectural lines, herringbone timber "
         "floor, warm pendant lighting, no people, professional medical education feel."),
    ],

    'consultation': [
        ("Elegant private medical consulting room, a polished timber desk with a soft brass desk "
         "lamp, deep leather chair, framed credentials on a white wall, afternoon light through "
         "timber venetian blinds, no people, prestigious and trustworthy."),
        ("Soft-focus reception area of a luxury private medical practice, fresh white orchids at "
         "the desk, polished stone floors reflecting overhead lighting, no people, welcoming "
         "and calm."),
        ("Two deep leather chairs facing each other across a low timber table in a beautifully lit "
         "consulting room, sunlight from a high window, no people, suggests a warm and considered "
         "professional conversation."),
    ],

    'pain': [
        ("Person (anonymous, back to camera only) seated in a deep armchair looking out through "
         "large windows at a sunlit private garden, contemplative and quiet mood, warm interior "
         "light, no clinical elements, empathetic and still."),
        ("Warm morning domestic scene, a comfortable sofa with soft linen cushions, a single hand "
         "resting gently on a knee (anonymous, no face, cropped), soft amber side light, intimate "
         "and empathetic."),
        ("Peaceful private garden seen through a slightly open door, warm soft focus bokeh of "
         "lawn and flowers, a sense of a better life waiting ahead, no people directly visible."),
    ],

    'surgery': [
        ("State-of-the-art private hospital building exterior at dusk, warm amber lights glowing "
         "from every window, clean architectural lines against a deep blue sky, no people, "
         "premium and reassuring."),
        ("Long private hospital corridor at dawn, polished floors reflecting overhead ambient "
         "light, doors receding into warm soft focus, pristine and calm, no people, premium "
         "private healthcare."),
        ("Private hospital garden courtyard at golden hour, mature frangipani tree, warm light "
         "on stone paths, calm and healing atmosphere, no people."),
    ],

    'implant': [
        ("Macro photograph of a precision-engineered titanium hip joint implant component on a "
         "clean white surface, soft studio lighting creating beautiful catchlights on the metal "
         "surface, product photography quality, no people."),
        ("Abstract macro of polished medical-grade titanium alloy, warm studio light creating "
         "soft gradients and reflections across a curved surface, sophisticated material "
         "photography, no people."),
        ("3D render of a hip joint in anatomical cross-section, dark background, clinical "
         "illustration quality, deep teal and silver tones, educational and precise."),
    ],

    'insurance': [
        ("Professional desk scene — an elegant leather-bound notebook open to a clean page, "
         "quality fountain pen, reading glasses, soft brass desk lamp, warm afternoon light, "
         "no people, sophisticated and trustworthy."),
        ("Warm private office interior, natural light through plantation shutters, a clean timber "
         "desk with a single neat folder open, aspirational professional setting, no people."),
    ],

    'ageing': [
        ("Mature couple (anonymous, backs to camera, faces not visible) walking hand-in-hand "
         "along a tree-lined avenue in the Adelaide Botanic Gardens at golden hour, warm light "
         "filtering through canopy, joyful mobility and independence."),
        ("Person (anonymous, back to camera, small in frame) hiking a gentle coastal trail with "
         "a scenic blue water vista ahead, warm golden light, sense of freedom and active ageing."),
        ("Group of older adults (anonymous, backs to camera, small in frame) doing tai chi in "
         "a park at dawn, soft morning mist through eucalyptus trees, peaceful and active."),
    ],

    'hospital': [
        ("Eastwood Private Hospital or premium private hospital exterior at dusk, warm amber "
         "lights behind glass, clean contemporary architecture, no people, premium private "
         "healthcare, aspirational."),
        ("Luxury private hospital lobby, soaring glass atrium with natural timber accents, "
         "warm ambient lighting, a single arrangement of fresh white flowers at the reception "
         "desk, no people, premium."),
        ("Private hospital room with a large window overlooking a healing garden, clean white "
         "linen, warm afternoon light on timber flooring, calm and reassuring, no people."),
    ],

    'gp': [
        ("General practitioner's private consulting room, warm timber desk, framed credentials "
         "on cream walls, afternoon light through timber venetian blinds, no people, trusted "
         "and professional."),
        ("Medical education seminar room, rows of empty upholstered chairs, a softly lit "
         "projected slide dimly visible, warm ambient lighting, professional CPD context, "
         "no people."),
    ],

    'adelaide': [
        ("Adelaide Hills vineyard at golden hour, rolling green hills, warm amber light on "
         "rows of vines, distinctly South Australian landscape, no people."),
        ("Glenelg Beach at sunrise, warm amber and rose-gold reflections on calm water, "
         "gentle waves, peaceful and South Australian, no people."),
        ("North Terrace, Adelaide, at dusk, warm heritage sandstone architecture lit "
         "beautifully against a deep blue sky, aspirational and locally resonant, no people."),
    ],

    'abstract': [
        ("Rich warm amber and deep gold abstract bokeh — large glowing circles dissolving "
         "softly out of focus against a near-black background, luxurious tonal depth, "
         "purely atmospheric, no people."),
        ("Deep navy-to-midnight-blue gradient, subtle soft points of light dissolving into "
         "darkness, luxurious and sophisticated, purely abstract, no people."),
        ("Soft rose-gold and warm cream bokeh, gentle glowing circles dissolving out of "
         "focus, calming and premium, no people."),
        ("Rich deep teal and navy abstract with subtle light rays radiating from the centre, "
         "sophisticated and atmospheric, no people."),
    ],
}


# ══════════════════════════════════════════════════════════════════════════════
# CLASSIFICATION
# ══════════════════════════════════════════════════════════════════════════════

def classify(topic, platform):
    t = topic.lower()
    p = platform.lower()

    if any(w in t for w in ['physio', 'rehabilitation', 'rehab']):
        return 'physio'
    if any(w in t for w in ['prehab', 'prehabilitation']):
        return 'exercise'
    if any(w in t for w in ['recover', 'post-op', 'post op', 'after surgery', 'returning',
                              'driving', 'milestone', 'discharge', 'icing', 'ice ',
                              'range before', 'walking after', 'home after']):
        return 'recovery'
    if any(w in t for w in ['exercise', 'walk', 'cycling', 'swim', 'sport', 'active ', 'fitness']):
        return 'exercise'
    if any(w in t for w in ['supplement', 'medication', 'drug', 'injection', 'cortisone',
                              'hyaluronic', 'nsaid', 'painkiller', 'script', 'prescri']):
        return 'supplements'
    if any(w in t for w in ['implant', 'prosthes', 'titanium', 'cemented', 'bearing',
                              'component', 'revision', 'patient specific', 'robotic',
                              'navigation', 'technology', 'computer assist']):
        return 'implant'
    if any(w in t for w in ['evidence', 'research', 'registry', 'aoanjrr', 'outcome',
                              'study', 'trial', 'audit', 'statistic', 'satisfaction',
                              'survival', 'revision rate', 'data ']):
        return 'research'
    if any(w in t for w in ['insurance', 'cover', 'gold hospital', 'cost', 'fund',
                              'financial', 'gap ', 'out of pocket', 'rebate']):
        return 'insurance'
    if any(w in t for w in ['pain', 'arthritis', 'osteoarthr', 'limp', 'stiff knee',
                              'stiff hip', 'waiting', 'when to', 'deciding']):
        return 'pain'
    if any(w in t for w in ['surgery', 'surgical', 'operation', 'theatre', 'anaesth',
                              'blood loss', 'transfusion', 'incision', 'approach',
                              'anterior ', 'posterior ', 'lateral ', 'nerve', 'complication',
                              'infection', 'eras', 'frailty', 'blood clot', 'dvt']):
        return 'surgery'
    if any(w in t for w in ['hospital', 'facility', 'private hospital', 'eastwood',
                              'inpatient', 'outpatient', 'day surgery', 'same-day']):
        return 'hospital'
    if any(w in t for w in ['weight', 'bmi', 'obese', 'obesity', 'overweight']):
        return 'wellness'
    if any(w in t for w in ['age', 'young ', 'older ', 'elderly', 'senior', 'ageing',
                              'aging', 'lifespan', 'longevity']):
        return 'ageing'
    if any(w in t for w in ['gp ', 'general pract', 'referral', 'primary care']):
        return 'gp'
    if any(w in t for w in ['consult', 'appointment', 'specialist', 'first visit',
                              'what to expect', 'how to prepare', 'what to bring',
                              'choose your', 'choosing']):
        return 'consultation'
    if any(w in t for w in ['adelaide', 'south australia', 'regional', 'rural']):
        return 'adelaide'
    if 'hip' in t and 'knee' not in t:
        return 'hip'
    if 'knee' in t and 'hip' not in t:
        return 'knee'
    if p == 'linkedin':
        return 'research'
    return 'abstract'


def get_background(pid, topic, platform):
    category = classify(topic, platform)
    options  = SCENES.get(category, SCENES['abstract'])
    return options[pid % len(options)]


# ══════════════════════════════════════════════════════════════════════════════
# PLATFORM PROMPT BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

BRAND_FOOTER = (
    "Overall design: sophisticated, warm, beautiful, trustworthy — looks like it was "
    "created by a senior art director for a premium private surgical practice."
)

PANEL_COMMON = (
    "Deep navy panel hex #16233A with gently rounded top corners. "
    "The panel does NOT touch the image edges — it floats inset with a small margin on all sides, "
    "creating breathing room. "
    "A soft diffused radiant glow of deep indigo-blue light emanates from BEHIND the panel and "
    "spreads outward beyond all four panel edges into the surrounding area — like the panel is "
    "gently backlit, hovering above the photograph. Subtle luxury, not neon — refined and elegant. "
    "Neue Montreal typeface throughout. "
    "All text left-aligned inside the panel, generous left padding. "
)

TYPOGRAPHY = (
    "LINE 1 — small uppercase, wide letter-spacing, colour #A2B9D8 (soft blue): ORTHOPAEDICS 360. "
    "LINE 2 — large bold white, wraps naturally across lines as needed, render EVERY word, "
    "absolutely no truncation: {topic}. "
    "LINE 3 — small regular white: Dr Chien-Wen Liew | Orthopaedics 360, Adelaide. "
)

WORDMARK = (
    "WORDMARK: the letters (DR) in small light-weight white, immediately followed by LIEW in "
    "significantly larger extra-bold white — size contrast is intentional and prominent, "
    "matching a luxury brand wordmark style. "
)


def instagram_prompt(pid, topic, background):
    return (
        "Luxury social media post, 1:1 square format, world-class creative agency, "
        "premium private orthopaedic surgical practice, Adelaide, Australia. "
        f"PHOTOGRAPHY — upper 58%: {background} "
        "Shallow depth of field, background softly blurred, editorial and cinematic. "
        "FLOATING PANEL — lower 42%: "
        + PANEL_COMMON
        + TYPOGRAPHY.format(topic=topic)
        + "BOTTOM ROW: LEFT — " + WORDMARK
        + "RIGHT — drchienwenliew.com.au in bold white, medium-large, prominently legible. "
        + BRAND_FOOTER
    )


def linkedin_prompt(pid, topic, background):
    return (
        "Luxury social media post, 16:9 wide format, world-class creative agency, "
        "premium private orthopaedic surgical practice, Adelaide, Australia. "
        "SPLIT LAYOUT: floating panel left 42%, editorial photography right 58%. "
        f"PHOTOGRAPHY — right 58%: {background} "
        "Shallow depth of field, rich cinematic tonal quality, professional and aspirational. "
        "FLOATING PANEL — left 42%: "
        "Deep navy panel hex #16233A with gently rounded right corners (top-right and bottom-right). "
        "Panel does NOT touch top, left, or bottom edges — floats inset with a small margin. "
        "A soft diffused radiant glow of deep indigo-blue light emanates from BEHIND the panel, "
        "spreading outward from the right edge into the photography side and all other edges — "
        "panel appears gently backlit and floating. Subtle luxury. "
        "Neue Montreal typeface throughout. All text left-aligned, generous padding. "
        + TYPOGRAPHY.format(topic=topic)
        + "BOTTOM ROW: LEFT — " + WORDMARK
        + "RIGHT — drchienwenliew.com.au bold white, clearly legible. "
        "Overall feel: authoritative, refined, premium medical professional — belongs in a "
        "premium hospital's annual report or a leading medical journal advertisement. "
        + BRAND_FOOTER
    )


def facebook_prompt(pid, topic, background):
    return (
        "Luxury social media post, 1:1 square format, world-class creative agency, "
        "premium private orthopaedic surgical practice, Adelaide, Australia. "
        f"PHOTOGRAPHY — upper 58%: {background} "
        "Shallow depth of field, editorial and cinematic, warm and approachable. "
        "FLOATING PANEL — lower 42%: "
        + PANEL_COMMON
        + TYPOGRAPHY.format(topic=topic)
        + "LINE 4 — drchienwenliew.com.au in bold white, medium-large, prominently legible. "
        "BOTTOM ROW: LEFT — " + WORDMARK
        + "RIGHT — IMPORTANT: leave a clean circular reserved area approximately 110 pixels "
        "diameter in the bottom-right corner of the panel, inside the panel's bottom-right area. "
        "Inside this circle: solid deep navy #16233A, no text, no design elements, no decoration. "
        "A thin subtle light border ring around the circle edge is optional for elegance. "
        "This circle will have a circular portrait photograph of the surgeon composited here later. "
        "No text from any line should overlap with or extend into this reserved circle area. "
        + BRAND_FOOTER
    )


# ══════════════════════════════════════════════════════════════════════════════
# CAROUSEL PROMPT BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def carousel_slide_prompt(pid, topic, platform, slide_num, background):
    """Single carousel dark slide — returns one prompt segment."""
    if platform.lower() == 'linkedin':
        return (
            f"Luxury 16:9 carousel slide {slide_num}, premium orthopaedic practice. "
            f"PHOTO right 58%: {background} Cinematic, editorial. "
            f"PANEL left 42%: deep navy #16233A rounded right corners, floating, "
            f"soft indigo backglow behind panel. Neue Montreal. "
            f"Small #A2B9D8 ORTHOPAEDICS 360 | Large bold white key point {slide_num} about '{topic}', "
            f"no truncation | Small white Dr Chien-Wen Liew | Orthopaedics 360 Adelaide. "
            f"Bottom: small (DR) + large bold LIEW left | drchienwenliew.com.au right."
        )
    else:
        return (
            f"Luxury 1:1 carousel slide {slide_num}, premium orthopaedic practice. "
            f"PHOTO top 58%: {background} Shallow DOF, cinematic, warm. "
            f"PANEL bottom 42%: deep navy #16233A rounded top corners, floating inset, "
            f"soft indigo glow from behind panel edges. Neue Montreal. "
            f"Small #A2B9D8 ORTHOPAEDICS 360 | Large bold white key point {slide_num} about '{topic}', "
            f"no truncation | Small white Dr Chien-Wen Liew. "
            f"Bottom: small (DR) + large bold LIEW left | drchienwenliew.com.au right."
        )


def carousel_prompts(pid, topic, platform):
    """Returns pipe-separated string of 4 slide prompts."""
    dark_slides = [2, 4, 6, 8]
    segments = []
    for i, slide_num in enumerate(dark_slides):
        # Rotate through different backgrounds for each slide
        bg_category = classify(topic, platform)
        options = SCENES.get(bg_category, SCENES['abstract'])
        background = options[(pid + i) % len(options)]
        seg = carousel_slide_prompt(pid, topic, platform, slide_num, background)
        segments.append(seg)
    return ' | '.join(segments)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print(f"\nLoading {XLSX.name}...")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb['Posts Schedule']

    headers = [cell.value for cell in ws[1]]
    col     = {h: i for i, h in enumerate(headers)}   # 0-based

    updated = skipped = carousel_count = 0

    for row in ws.iter_rows(min_row=2):
        pid_cell = row[col['ID']]
        if pid_cell.value is None:
            continue
        try:
            pid = int(pid_cell.value)
        except (TypeError, ValueError):
            continue
        if pid < ACTIVE_FROM_ID:
            skipped += 1
            continue

        platform = str(row[col['Platform']].value or '').strip()
        topic    = str(row[col['Topic']].value or '').strip()
        fmt      = str(row[col['Format']].value or '').strip().lower()
        bg       = get_background(pid, topic, platform)

        is_carousel = 'carousel' in fmt or 'infographic' in fmt

        if is_carousel:
            prompt = carousel_prompts(pid, topic, platform)
            carousel_count += 1
        elif platform.lower() == 'linkedin':
            prompt = linkedin_prompt(pid, topic, bg)
        elif platform.lower() == 'facebook':
            prompt = facebook_prompt(pid, topic, bg)
        else:
            prompt = instagram_prompt(pid, topic, bg)

        row[col['Kie.ai Image Description']].value = prompt
        updated += 1

        plat_label = {'Instagram': 'IG', 'LinkedIn': 'LI', 'Facebook': 'FB'}.get(platform, platform[:2])
        fmt_label  = '(carousel)' if is_carousel else ''
        print(f"  {pid:3d}  {plat_label}  {topic[:60]}{fmt_label}")

    wb.save(XLSX)

    print()
    print('═' * 58)
    print(f'  ✅  {updated} posts updated  ({carousel_count} carousels)')
    print(f'      {skipped} archived posts skipped')
    print(f'      Saved → {XLSX.name}')
    print()
    print('  Next steps:')
    print('  1. Regenerate Posts from Excel.command')
    print('  2. Generate Kie.ai Images.command   (or --from-date)')
    print('  3. Push Images to GitHub.command')
    print('═' * 58)
    print()


if __name__ == '__main__':
    main()
